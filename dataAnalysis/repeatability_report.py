#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Repeatability report from diagnosis_log.csv -> Repeatability_test.xlsx.
Uses only Channel + six reply columns (s1/s2/s3 pd/opm); all pre_* columns ignored.
"""

from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

INPUT_CSV = "diagnosis_log.csv"
OUTPUT_XLSX = "Repeatability_test.xlsx"
TITLE_ROW1 = "\u6302\u673a\u6c47\u603b\u6570\u636e"

WAVELENGTHS: List[Tuple[str, str, str, str]] = [
    ("SFP_1550", "1550", "s1_pd_reply", "s1_opm_reply"),
    ("SFP_1310", "1310", "s2_pd_reply", "s2_opm_reply"),
    ("Laser_1310", "Laser", "s3_pd_reply", "s3_opm_reply"),
]

REQUIRED_COLUMNS = ["Channel"] + [c for _, _, pd, pm in WAVELENGTHS for c in (pd, pm)]

INVALID_RAW = frozenset({-6000, -999900, 999900})
RAW_ABS_MAX = 900000

FILL_HEADER = PatternFill("solid", fgColor="00B0F0")
FILL_SUMMARY = PatternFill("solid", fgColor="FF0000")
FILL_RESULT_COL = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_SUMMARY = Font(bold=True, color="FFFFFF", size=10)
FONT_BOLD = Font(bold=True, size=10)


@dataclass
class WlStats:
    il_max: Optional[float] = None
    il_min: Optional[float] = None
    il_result: Optional[float] = None
    sample_count: int = 0


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def message_box(title: str, text: str, error: bool = False) -> None:
    try:
        import ctypes

        flags = 0x10 if error else 0x40
        ctypes.windll.user32.MessageBoxW(0, text, title, flags)
    except Exception:
        print(f"{title}: {text}", file=sys.stderr)


def parse_int(raw: str) -> Optional[int]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def is_valid_raw(v: int) -> bool:
    if v in INVALID_RAW:
        return False
    if abs(v) >= RAW_ABS_MAX:
        return False
    return True


def compute_il(pd_raw: int, pm_raw: int) -> float:
    return pm_raw / 10000.0 - pd_raw / 100.0


def channel_sort_key(ch: str) -> Tuple[int, str]:
    m = re.search(r"(\d+)", ch or "")
    return (int(m.group(1)) if m else 0, ch.upper())


def format_channel_display(ch: str) -> str:
    m = re.search(r"(\d+)", ch or "")
    if m:
        return f"ch{m.group(1)}"
    return (ch or "").lower()


def load_and_aggregate(csv_path: Path) -> Dict[str, Dict[str, WlStats]]:
    """Group IL samples by channel and wavelength label (SFP_1550, ...)."""
    il_lists: Dict[str, Dict[str, List[float]]] = defaultdict(
        lambda: {wl[0]: [] for wl in WAVELENGTHS}
    )

    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row.")
        missing = [c for c in REQUIRED_COLUMNS if c not in reader.fieldnames]
        if missing:
            raise ValueError(
                "CSV missing required columns: " + ", ".join(missing)
            )

        for row in reader:
            ch = (row.get("Channel") or "").strip()
            if not ch:
                continue
            for wl_label, _short, pd_col, pm_col in WAVELENGTHS:
                pd_v = parse_int(row.get(pd_col))
                pm_v = parse_int(row.get(pm_col))
                if pd_v is None or pm_v is None:
                    continue
                if not is_valid_raw(pd_v) or not is_valid_raw(pm_v):
                    continue
                il_lists[ch][wl_label].append(compute_il(pd_v, pm_v))

    result: Dict[str, Dict[str, WlStats]] = {}
    for ch, per_wl in il_lists.items():
        result[ch] = {}
        for wl_label, _short, _, _ in WAVELENGTHS:
            vals = per_wl.get(wl_label) or []
            st = WlStats(sample_count=len(vals))
            if vals:
                st.il_max = max(vals)
                st.il_min = min(vals)
                st.il_result = st.il_max - st.il_min
            result[ch][wl_label] = st
    return result


def write_xlsx(path: Path, agg: Dict[str, Dict[str, WlStats]]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Repeatability"

    ws.merge_cells("A1:J1")
    c1 = ws["A1"]
    c1.value = TITLE_ROW1
    c1.fill = FILL_HEADER
    c1.font = FONT_HEADER
    c1.alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = "IL = Power_PM - Power_PD"
    ws["A2"].font = FONT_BOLD

    ws["B3"] = "SFP_1550"
    ws["E3"] = "SFP_1310"
    ws["H3"] = "Laser_1310"
    for col in (2, 5, 8):
        cell = ws.cell(row=3, column=col)
        cell.font = FONT_BOLD
        cell.alignment = Alignment(horizontal="center")

    headers = ["channel"]
    for _wl, short, _, _ in WAVELENGTHS:
        headers.extend([f"IL_Max {short}", f"IL_Min {short}", f"IL_result {short}"])
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = FONT_BOLD
        if "IL_result" in h:
            cell.fill = FILL_RESULT_COL

    result_col_indices = [4, 7, 10]
    for wi, (wl_label, _, _, _) in enumerate(WAVELENGTHS):
        mx = None
        for per_ch in agg.values():
            st = per_ch.get(wl_label)
            if st and st.il_result is not None:
                mx = st.il_result if mx is None else max(mx, st.il_result)
        col = result_col_indices[wi]
        cell = ws.cell(row=4, column=col)
        if mx is not None:
            cell.value = round(mx, 4)
        cell.fill = FILL_SUMMARY
        cell.font = FONT_SUMMARY
        cell.number_format = "0.0000"

    channels = sorted(agg.keys(), key=channel_sort_key)
    row = 6
    for ch in channels:
        ws.cell(row=row, column=1, value=format_channel_display(ch))
        col = 2
        for wl_label, _, _, _ in WAVELENGTHS:
            st = agg[ch].get(wl_label, WlStats())
            for val in (st.il_max, st.il_min, st.il_result):
                cell = ws.cell(row=row, column=col)
                if val is not None:
                    cell.value = round(val, 4)
                    cell.number_format = "0.0000"
                if col in result_col_indices:
                    cell.fill = FILL_RESULT_COL
                col += 1
        row += 1

    ws.column_dimensions["A"].width = 12
    for c in range(2, 11):
        ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(path)


def main() -> int:
    base = app_dir()
    csv_path = base / INPUT_CSV
    out_path = base / OUTPUT_XLSX

    if not csv_path.is_file():
        message_box(
            "Repeatability Report",
            f"Input file not found:\n{csv_path}\n\nPlace {INPUT_CSV} next to this program.",
            error=True,
        )
        return 1

    try:
        agg = load_and_aggregate(csv_path)
        if not agg:
            message_box(
                "Repeatability Report",
                "No valid data rows found in CSV (check Channel and reply columns).",
                error=True,
            )
            return 1
        write_xlsx(out_path, agg)
    except Exception as e:
        message_box("Repeatability Report", f"Failed:\n{e}", error=True)
        return 1

    n_ch = len(agg)
    message_box(
        "Repeatability Report",
        f"Generated successfully:\n{out_path}\n\nChannels: {n_ch}",
        error=False,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
